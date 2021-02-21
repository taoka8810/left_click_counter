#
# 0:00から23:59までに左クリックした回数をカウントし、Excelファイルに記録するプログラム
#

import ctypes, sys
import openpyxl as px
from datetime import datetime

wb=px.Workbook()
ws=wb.worksheets[0]
ws.cell(row=1, column=1).value=0
counter=0


while True:
    #23:58にプログラムを停止する    
    now=datetime.now()
    if now.hour==23:
        if now.minute==59:
            if now.second==40:
                sys.exit()
    #左クリックを検知、Excelファイルに記録
    if ctypes.windll.user32.GetAsyncKeyState(0x01)==0x8000:
        while True:
            if ctypes.windll.user32.GetAsyncKeyState(0x01)!=0x8000:
                counter+=1
                x=ws.max_row+1
                ws.cell(row=x, column=1).value=counter
                wb.save("record.xlsx")
                print("ファイルに書き込みました:"+str(counter)+"回目")
                break

        




  
