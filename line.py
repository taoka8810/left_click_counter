#
#click_counter.pyで記録したExcelファイルを読み込んで、左クリックした回数をLINEで通知するプログラム 
#

import json, os, urllib3
import openpyxl as px
from datetime import date
from linebot import LineBotApi
from linebot.models import TextSendMessage
from urllib3.exceptions import InsecureRequestWarning

urllib3.disable_warnings(InsecureRequestWarning)
today=date.today()
path_json=r"C:\Users\taoka\Desktop\program\python\main\left_click\info.json"
path_xlsx=os.path.join(r"C:\Users\taoka\Desktop\program\python\main\left_click\record.xlsx")

file_json=open(path_json, "r")
info=json.load(file_json)
wb=px.load_workbook(path_xlsx)
ws=wb.worksheets[0]
x=ws.max_row
n=ws.cell(row=x, column=1).value
text="本日左クリックをした回数："+str(n)+"回"

CHANNEL_ACCESS_TOKEN=info["CHANNEL_ACCESS_TOKEN"]
line_bot_api=LineBotApi(CHANNEL_ACCESS_TOKEN)

def main():
    USER_ID=info["USER_ID"]
    messages=TextSendMessage(text)
    line_bot_api.push_message(USER_ID, messages)

main()
os.remove("record.xlsx")



